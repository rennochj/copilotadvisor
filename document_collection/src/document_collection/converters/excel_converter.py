"""Excel to Markdown converter implementation."""

import zipfile
from pathlib import Path
from typing import Any

from document_collection.core.interfaces import DocumentConverter


class ExcelConverter(DocumentConverter):
    """Convert Excel documents to Markdown format with image/chart extraction."""

    def can_convert(self, file_path: Path) -> bool:
        """Check if this converter can handle Excel files."""
        return str(file_path).lower().endswith(".xlsx")

    def get_supported_formats(self) -> list[str]:
        """Return list of supported formats."""
        return ["xlsx"]

    async def convert(self, input_path: Path, output_path: Path, **kwargs: Any) -> Path:
        """Convert Excel document to Markdown using openpyxl with image/chart extraction."""
        try:
            from openpyxl import load_workbook

            # Read Excel workbook
            workbook = load_workbook(str(input_path), data_only=True)

            # Create images directory
            images_dir = output_path.parent / "images"
            images_dir.mkdir(parents=True, exist_ok=True)

            # Extract embedded images and charts from the workbook's media folder
            image_counter = 1
            image_mapping = {}  # Map original image names to new filenames

            try:
                # Access the workbook as a zip file to extract images/charts
                with zipfile.ZipFile(str(input_path), 'r') as xlsx_zip:
                    # Look for images and charts in the media directory
                    for file_info in xlsx_zip.filelist:
                        if file_info.filename.startswith('xl/media/'):
                            # Extract image/chart
                            image_data = xlsx_zip.read(file_info.filename)

                            # Determine file extension
                            original_name = Path(file_info.filename).name
                            file_ext = Path(file_info.filename).suffix.lower()
                            if not file_ext:
                                file_ext = '.png'  # Default

                            # Create new filename
                            new_filename = f"{input_path.stem}_image_{image_counter:03d}{file_ext}"
                            image_path = images_dir / new_filename

                            # Save image/chart
                            with open(image_path, 'wb') as img_file:
                                img_file.write(image_data)

                            # Map original to new filename
                            image_mapping[original_name] = new_filename
                            image_counter += 1

            except Exception:
                # Continue if image extraction fails
                pass

            # Convert workbook content to markdown
            content_parts = [f"# {input_path.stem}\n"]
            content_parts.append("Converted from Excel workbook\n")

            if image_counter > 1:
                content_parts.append(f"*This workbook contains {image_counter - 1} extracted images/charts stored in the `images/` directory.*\n")

            # Process worksheets
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                content_parts.append(f"\n## Sheet: {sheet_name}\n")

                # Find the actual data range (skip empty rows/columns)
                rows_with_data: list[list[Any]] = []
                for row_tuple in worksheet.iter_rows():
                    row_values = [cell.value for cell in row_tuple if cell.value is not None]
                    if row_values:  # Only include rows with data
                        rows_with_data.append([cell.value for cell in row_tuple])

                if rows_with_data:
                    # Determine the maximum number of columns with data
                    max_cols = max(len([val for val in row if val is not None]) for row in rows_with_data)

                    # Filter out completely empty columns
                    filtered_rows: list[list[Any]] = []
                    for row in rows_with_data:
                        # Pad row to max_cols length and ensure it's a list
                        row_list = list(row)[:max_cols]
                        padded_row = row_list + [None] * (max_cols - len(row_list))
                        filtered_rows.append(padded_row)

                    if filtered_rows:
                        content_parts.append("\n")
                        for i, row in enumerate(filtered_rows):
                            # Convert None values to empty strings and ensure all values are strings
                            row_cells = [str(cell) if cell is not None else "" for cell in row]
                            content_parts.append("| " + " | ".join(row_cells) + " |\n")

                            # Add header separator for first row
                            if i == 0:
                                content_parts.append("| " + " | ".join(["---"] * len(row_cells)) + " |\n")
                        content_parts.append("\n")
                else:
                    content_parts.append("*This sheet is empty*\n")

            # Add image references for any images found
            if image_mapping:
                content_parts.append("\n## Extracted Images and Charts\n\n")
                for i, (_original, new_filename) in enumerate(image_mapping.items(), 1):
                    content_parts.append(f"![Image/Chart {i}](images/{new_filename})\n\n")

            # Write markdown content
            markdown_content = "\n".join(content_parts)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(markdown_content)

            return output_path

        except ImportError:
            # Fallback if openpyxl is not available
            output_path.parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(f"# {input_path.stem}\n\nExcel conversion requires openpyxl library.\n")
            return output_path
        except Exception as e:
            # Create error file
            output_path.parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(f"# {input_path.stem}\n\nError converting Excel document: {str(e)}\n")
            return output_path
